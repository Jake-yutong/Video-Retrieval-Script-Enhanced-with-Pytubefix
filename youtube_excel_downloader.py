#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
YouTube Excel视频下载工具
功能：根据Excel文档下载YouTube视频，按编号命名，超长视频分段
"""

import os
import sys
import subprocess
import json
import re
from datetime import datetime as dt, timedelta
from pathlib import Path
from typing import Optional, List, Dict
import openpyxl

# 添加 deno 到 PATH（yt-dlp 需要）
DENO_PATH = "/Users/liyutong/.deno/bin"
if DENO_PATH not in os.environ.get('PATH', ''):
    os.environ['PATH'] = DENO_PATH + ":" + os.environ.get('PATH', '')


# ============ 配置 ============
OUTPUT_DIR = Path("/Volumes/T7 Shield/纪录片collection_1.9")
EXCEL_FILE = Path("/Volumes/T7 Shield/Tour-related Video Info.xlsx")

# 视频分段阈值（30分钟 = 1800秒）- 仅YouTube支持分段
SEGMENT_THRESHOLD_SECONDS = 30 * 60  # 30分钟
SEGMENT_DURATION_SECONDS = 10 * 60   # 每段10分钟

# 下载质量
VIDEO_QUALITY = "best[height<=360]"


def get_url_platform(url: str) -> str:
    """判断URL平台类型"""
    url_str = str(url).lower()
    if 'youtube.com' in url_str or 'youtu.be' in url_str:
        return 'youtube'
    elif 'bilibili.com' in url_str:
        return 'bilibili'
    elif 'rthk.hk' in url_str:
        return 'rthk'
    else:
        return 'other'


def parse_duration(duration) -> int:
    """解析时长，返回秒数"""
    if duration is None:
        return 0

    if isinstance(duration, dt):
        # datetime.datetime 或 datetime.time
        if hasattr(duration, 'hour'):
            return duration.hour * 3600 + duration.minute * 60 + duration.second
        return 0

    if isinstance(duration, timedelta):
        return int(duration.total_seconds())

    if isinstance(duration, (int, float)):
        return int(duration)

    duration_str = str(duration)

    # 匹配格式: "84mins", "23mins", "26mins", "46mins"
    mins_match = re.search(r'(\d+)\s*mins?', duration_str, re.IGNORECASE)
    if mins_match:
        return int(mins_match.group(1)) * 60

    # 匹配格式: "20:35" (分:秒)
    time_match = re.search(r'(\d+):(\d+)', duration_str)
    if time_match:
        minutes = int(time_match.group(1))
        seconds = int(time_match.group(2))
        return minutes * 60 + seconds

    return 0


def format_video_id(url: str) -> str:
    """从URL提取视频ID"""
    # YouTube URL patterns
    patterns = [
        r'(?:v=|/v/|youtu\.be/)([a-zA-Z0-9_-]{11})',
        r'watch\?v=([a-zA-Z0-9_-]{11})',
    ]

    for pattern in patterns:
        match = re.search(pattern, url)
        if match:
            return match.group(1)

    return None


def download_video_segment(url: str, output_name: str, output_dir: Path,
                          start_time: int = None, end_time: int = None,
                          platform: str = 'youtube') -> bool:
    """
    下载视频片段

    Args:
        url: 视频链接
        output_name: 输出文件名（不含扩展名）
        output_dir: 输出目录
        start_time: 开始时间（秒）- 仅YouTube支持
        end_time: 结束时间（秒）- 仅YouTube支持
        platform: 平台类型 (youtube/bilibili/rthk/other)
    """

    # 构建输出路径
    output_path = output_dir / f"{output_name}.mp4"

    # 如果视频已存在，跳过
    if output_path.exists():
        print(f"   ⏭️ 已存在: {output_name}.mp4")
        return True

    print(f"   📥 下载 [{platform}]: {output_name}")

    # 构建下载命令
    cmd = ['yt-dlp']

    # YouTube特殊参数
    if platform == 'youtube':
        cmd.extend(['--remote-components', 'ejs:github'])

    # 下载质量
    if platform in ('youtube', 'bilibili'):
        cmd.extend(['-f', VIDEO_QUALITY])

    cmd.extend([
        '--no-playlist',
        '--no-check-certificate',
        '--merge-output-format', 'mp4',
        '-o', str(output_path),
    ])

    # 字幕支持（YouTube和Bilibili）
    if platform in ('youtube', 'bilibili'):
        cmd.extend([
            '--write-subs',
            '--sub-lang', 'en,zh-Hans,zh-Hant,zh',
            '--convert-subs', 'vtt',
        ])

    # 添加时间片段参数（仅YouTube支持）
    if platform == 'youtube' and start_time is not None:
        cmd.extend(['--download-sections', f'*{start_time}-{end_time}' if end_time else f'*{start_time}-'])

    # 添加URL
    cmd.append(url)

    try:
        result = subprocess.run(
            cmd,
            capture_output=True,
            text=True,
            timeout=900  # 15分钟超时（长视频）
        )

        if result.returncode == 0:
            print(f"   ✅ 完成: {output_name}")
            return True
        else:
            print(f"   ❌ 失败: {output_name}")
            if result.stderr:
                stderr_lower = result.stderr.lower()
                if 'private video' in stderr_lower or 'privat' in stderr_lower:
                    print(f"      原因: 私有视频")
                elif 'is not available' in stderr_lower:
                    print(f"      原因: 视频不可用")
                elif 'login required' in stderr_lower:
                    print(f"      原因: 需要登录")
                else:
                    print(f"      错误: {result.stderr[-200:]}")
            return False

    except subprocess.TimeoutExpired:
        print(f"   ⏰ 超时: {output_name}")
        return False
    except Exception as e:
        print(f"   ❌ 错误: {output_name} - {e}")
        return False


def process_video(video_no: int, title: str, url: str, output_dir: Path) -> bool:
    """
    处理单个视频

    Args:
        video_no: 视频编号
        title: 视频标题
        url: 视频链接
        output_dir: 输出目录

    Returns:
        是否成功
    """
    # 格式化编号 (1 -> "001")
    video_id_str = f"{video_no:03d}"

    # 判断平台
    platform = get_url_platform(url)
    print(f"\n[{video_id_str}] {title[:45]}... [{platform}]")

    # 获取视频时长（YouTube和Bilibili支持）
    duration_seconds = 0
    if platform in ('youtube', 'bilibili'):
        cmd = [
            'yt-dlp',
            '--print', '%(duration)s',
            '--no-download',
            url
        ]
        try:
            result = subprocess.run(cmd, capture_output=True, text=True, timeout=30)
            if result.returncode == 0 and result.stdout.strip():
                duration_seconds = int(result.stdout.strip())
        except:
            duration_seconds = 0

        print(f"   时长: {duration_seconds // 60}分{duration_seconds % 60}秒")

    # 确保输出目录存在
    output_dir.mkdir(parents=True, exist_ok=True)

    # YouTube长视频需要分段（Bilibili/RTHK不支持分段下载）
    if platform == 'youtube' and duration_seconds > SEGMENT_THRESHOLD_SECONDS:
        print(f"   ✂️ 需要分段 (>{SEGMENT_THRESHOLD_SECONDS // 60}分钟)")

        num_segments = (duration_seconds + SEGMENT_DURATION_SECONDS - 1) // SEGMENT_DURATION_SECONDS

        success_count = 0
        for seg_idx in range(num_segments):
            seg_start = seg_idx * SEGMENT_DURATION_SECONDS
            seg_end = min((seg_idx + 1) * SEGMENT_DURATION_SECONDS, duration_seconds)

            # 格式化分段名称: 001_01, 001_02, etc.
            seg_name = f"{video_id_str}_{seg_idx + 1:02d}"

            if download_video_segment(url, seg_name, output_dir, seg_start, seg_end, platform):
                success_count += 1

        print(f"   分段完成: {success_count}/{num_segments}")
        return success_count > 0
    else:
        # 单段下载
        return download_video_segment(url, video_id_str, output_dir, platform=platform)


def read_excel_videos(excel_path: Path) -> List[Dict]:
    """读取Excel文件中的视频信息（支持所有平台）"""
    videos = []

    wb = openpyxl.load_workbook(excel_path)
    sheet = wb.active

    print(f"📊 读取Excel文件: {excel_path.name}")
    print(f"   总行数: {sheet.max_row}")

    youtube_count = 0
    bilibili_count = 0
    rthk_count = 0
    other_count = 0

    for i, row in enumerate(sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True), 2):
        no = row[0]  # 列A: 编号
        title = row[1]  # 列B: 标题
        url = row[6]  # 列G: URL

        if not url:
            continue

        url_str = str(url)
        platform = get_url_platform(url)

        # 跳过YouTube playlist
        if platform == 'youtube' and '&list=' in url_str:
            continue

        videos.append({
            'no': int(no) if no else len(videos) + 1,
            'title': str(title).strip() if title else f"Video_{len(videos) + 1}",
            'url': url_str,
            'platform': platform
        })

        if platform == 'youtube':
            youtube_count += 1
        elif platform == 'bilibili':
            bilibili_count += 1
        elif platform == 'rthk':
            rthk_count += 1
        else:
            other_count += 1

    print(f"   视频统计:")
    print(f"   - YouTube: {youtube_count}")
    print(f"   - Bilibili: {bilibili_count}")
    print(f"   - RTHK: {rthk_count}")
    print(f"   - Other: {other_count}")
    print(f"   - 总计: {len(videos)} 个视频\n")

    wb.close()
    return videos


def main():
    """主函数"""
    print("="*60)
    print("   Excel 视频下载工具（支持YouTube/B站/RTHK）")
    print("   功能: 按编号下载 | YouTube超长分段 | 360p | 带字幕")
    print("="*60)

    # 确保输出目录存在
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    print(f"\n📁 输出目录: {OUTPUT_DIR}")

    # 读取Excel
    if not EXCEL_FILE.exists():
        print(f"❌ Excel文件不存在: {EXCEL_FILE}")
        sys.exit(1)

    videos = read_excel_videos(EXCEL_FILE)

    if not videos:
        print("❌ 未找到视频链接")
        sys.exit(1)

    # 下载每个视频
    success_count = 0
    fail_count = 0
    skip_count = 0

    print("🚀 开始下载...\n")

    for i, video in enumerate(videos, 1):
        print(f"[{i}/{len(videos)}]", "="*50)

        # 检查是否已存在（跳过已下载的）
        video_id_str = f"{video['no']:03d}"
        output_path = OUTPUT_DIR / f"{video_id_str}.mp4"
        if output_path.exists():
            print(f"   ⏭️ 已存在: {video_id_str}.mp4")
            skip_count += 1
            continue

        if process_video(video['no'], video['title'], video['url'], OUTPUT_DIR):
            success_count += 1
        else:
            fail_count += 1

    # 摘要
    print("\n" + "="*60)
    print("   下载完成！")
    print("="*60)
    print(f"   成功: {success_count}")
    print(f"   失败: {fail_count}")
    print(f"   跳过: {skip_count}")
    print(f"   总计: {len(videos)}")
    print(f"   输出目录: {OUTPUT_DIR}")
    print("="*60)


if __name__ == "__main__":
    main()
